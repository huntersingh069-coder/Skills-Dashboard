import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import datetime

wb = openpyxl.Workbook(https://kentechcloud-my.sharepoint.com/:x:/g/personal/nayaabshaad_ansari_kentplc_com/IQAYmbnAbVdGSr-o-cJHxgQLAd63DirBYMoj7EZZz7TrjgY?e=RifmeN)

# ---- palette ----
HEADER_FILL = PatternFill("solid", fgColor="1F2E3D")
INPUT_FILL = PatternFill("solid", fgColor="FFF3CC")     # yellow = manual entry
FORMULA_FILL = PatternFill("solid", fgColor="E9EEF2")   # light grey = auto-calculated
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F2E3D")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
thin = Side(style="thin", color="C9D2D8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ============================================================= LEVEL / SECTION DEFINITIONS
# The "label" text below must match EXACTLY what the dashboard website expects
# (it is used as the dropdown validation list on the Profile sheet, and the website
# maps each label to its corresponding Skills_* sheet). Do not edit the labels.
SECTIONS = [
    {
        "label": "GET/DET (0-1 yrs)",
        "sheet": "Skills_GETDET",
        "skills": [
            "Basic Electrical Theory & Circuit Fundamentals",
            "Reading Electrical Drawings (SLD, Cable Schedules, Layouts)",
            "Awareness of IEC/IEEE/API Standards",
            "Basic Cable Sizing & Load List Preparation",
            "Documentation & MS Office Practices",
            "Site Safety Awareness (HSE Induction, PTW Basics)",
            "Basic AutoCAD / Drafting Support",
            "Equipment Datasheet Familiarization",
            "Assisting Vendor Document Reviews",
            "Basic Earthing & Lighting Concepts",
        ],
    },
    {
        "label": "Engineer (1-7 yrs)",
        "sheet": "Skills_Engineer",
        "skills": [
            "Power System Studies (Load Flow / Short Circuit Basics)",
            "Hazardous Area Classification Application",
            "Electrical Load Calculations & Sizing",
            "Cable Sizing & Voltage Drop Calculations",
            "Lighting Design & Calculations",
            "Earthing & Lightning Protection Design",
            "Motor & Drive System Engineering",
            "Transformer Sizing & Protection Basics",
            "Switchgear & MCC Specification",
            "Vendor Document Review & Technical Bid Evaluation",
            "Codes & Standards Compliance (IEC/IEEE/API/NEC)",
            "Cathodic Protection Systems Basics",
        ],
    },
    {
        "label": "Sr. Engineer (7-14 yrs)",
        "sheet": "Skills_SrEngineer",
        "skills": [
            "Advanced Power System Studies (Arc Flash, Harmonics)",
            "Protection & Relay Coordination Philosophy",
            "Electrical Design Basis & Philosophy Development",
            "HAZOP / SIL Participation (Electrical Scope)",
            "Fire & Gas Detection Systems (Electrical Interface)",
            "Instrumentation & Electrical Interface Coordination",
            "Cathodic Protection System Design",
            "Package Vendor Engineering Coordination",
            "Technical Query & RFI Resolution",
            "Junior Engineer Mentoring & Technical Review",
            "Project Electrical Scope Estimation",
        ],
    },
    {
        "label": "Principal Engineer (14-18 yrs)",
        "sheet": "Skills_PrincipalEngineer",
        "skills": [
            "Electrical Engineering Design Basis Ownership",
            "Multi-Discipline Technical Coordination",
            "FEED & Detailed Engineering Leadership",
            "Client & Third-Party Technical Interface",
            "Technical Risk Assessment & Mitigation Planning",
            "Electrical Philosophy for Brownfield/Greenfield Projects",
            "Engineering Standards Development & Gap Analysis",
            "Value Engineering & Cost Optimization",
            "Mentoring Sr. Engineers & Career Development",
            "Root Cause Analysis for Electrical System Failures",
        ],
    },
    {
        "label": "Dy. Chief/Chief Engineer (18+ yrs)",
        "sheet": "Skills_ChiefEngineer",
        "skills": [
            "Electrical Engineering Strategy & Governance",
            "Technology Roadmap & Digitalization Initiatives",
            "Cross-Project Technical Assurance",
            "Engineering Centre of Excellence Leadership",
            "Client Relationship & Business Development Support",
            "Organizational Competency Development Framework",
            "Major Project Technical Authority Sign-off",
            "Industry Standards Committee Participation",
            "Crisis / Incident Technical Leadership",
            "Succession Planning & Talent Pipeline",
        ],
    },
]

SOFTWARE = [
    "AutoCAD 2D",
    "ETAP",
    "SKM",
    "ORACLE",
    "Advanced Excel",
    "SPEL",
    "Navisworks Freedom",
    "DiaLUX EVO / DiaLUX 4.13",
]

# one realistic sample employee per level, so every section has a working demo row
SAMPLES = [
    {"name": "Rohan Mehta",      "position": "Graduate Engineer Trainee",         "section": 0, "doj": datetime.date(2025, 8, 1),  "overall": 0.9,
     "skills": [3, 2, 2, 2, 4, 3, 2, 2, 2, 1], "sw": [2, 0, 0, 1, 3, 0, 0, 0]},
    {"name": "Priya Sharma",     "position": "Electrical Engineer",               "section": 1, "doj": datetime.date(2021, 6, 10), "overall": 5.0,
     "skills": [4, 3, 4, 5, 4, 3, 3, 3, 4, 3, 4, 2], "sw": [4, 2, 1, 2, 4, 1, 2, 3]},
    {"name": "Ahmed Al-Mansoori","position": "Senior Electrical Engineer",        "section": 2, "doj": datetime.date(2019, 3, 15), "overall": 9.8,
     "skills": [4, 4, 4, 4, 3, 4, 3, 3, 4, 3, 4], "sw": [5, 4, 3, 2, 4, 3, 2, 3]},
    {"name": "Vikram Nair",      "position": "Principal Electrical Engineer",     "section": 3, "doj": datetime.date(2013, 1, 20), "overall": 16.2,
     "skills": [5, 4, 5, 4, 4, 4, 4, 3, 4, 5], "sw": [5, 5, 4, 3, 5, 3, 2, 3]},
    {"name": "Suresh Iyer",      "position": "Deputy Chief Electrical Engineer",  "section": 4, "doj": datetime.date(2006, 9, 5),  "overall": 22.5,
     "skills": [5, 4, 4, 5, 4, 4, 5, 4, 4, 5], "sw": [4, 4, 3, 4, 5, 2, 2, 2]},
]

RATING_LEVELS = [
    (0, "No exposure"), (1, "Basic awareness"), (2, "Working knowledge"),
    (3, "Competent, needs occasional guidance"), (4, "Proficient, works independently"),
    (5, "Expert / can train others"),
]

def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        cell.border = BORDER

# ============================================================= INSTRUCTIONS
ws = wb.active
ws.title = "Instructions"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 100
ws.column_dimensions["C"].width = 55

ws["B2"] = "Personnel Skills Matrix — Data Template"
ws["B2"].font = TITLE_FONT

lines = [
    ("How this workbook is structured", True),
    ("This workbook feeds the dashboard website. Keep sheet names and column headers exactly as they are — "
     "the website reads them by name.", False),
    ("", False),
    ("1. Profile sheet", True),
    ("One row per employee. Fill in Name, Position, Section (choose from the dropdown — this determines which "
     "level/band the employee belongs to and which skillset applies to them), and Date of Joining (yellow cells). "
     "\"Experience with this Organization\" is calculated automatically from the Date of Joining — do not type into it. "
     "\"Overall Experience\" is typed manually (e.g. 9.8, 8.4) since it includes experience from before this organization.", False),
    ("", False),
    ("2. The five Skills_* sheets", True),
    ("Each career level has its OWN skillset on its OWN sheet, because a GET/DET's skillset is very different from "
     "a Chief Engineer's:", False),
    ("   • Skills_GETDET  →  GET/DET (0-1 yrs)", False),
    ("   • Skills_Engineer  →  Engineer (1-7 yrs)", False),
    ("   • Skills_SrEngineer  →  Sr. Engineer (7-14 yrs)", False),
    ("   • Skills_PrincipalEngineer  →  Principal Engineer (14-18 yrs)", False),
    ("   • Skills_ChiefEngineer  →  Dy. Chief/Chief Engineer (18+ yrs)", False),
    ("Add each employee's row ONLY on the sheet matching their Section. One row per employee, one column per skill. "
     "Enter a Rating from 0 to 5 in each cell.", False),
    ("", False),
    ("3. Software sheet", True),
    ("Common across all levels. One row per employee, one column per software (8 columns). Enter a Rating from 0 to 5.", False),
    ("", False),
    ("Rating scale (0–5) used on all Skills_* sheets and the Software sheet", True),
    ("", False),
    ("__RATING_TABLE__", False),
    ("", False),
    ("Color legend", True),
    ("Yellow fill = type your value here.   Grey fill = calculated automatically, do not overwrite.", False),
    ("", False),
    ("Keeping names consistent", True),
    ("The Name value must be spelled identically on the Profile sheet, the matching Skills_* sheet, and the Software "
     "sheet — this is how the dashboard links a person's profile to their ratings.", False),
]

r = 4
for text, is_heading in lines:
    if text == "__RATING_TABLE__":
        ws.cell(row=r, column=2, value="Level").font = HEADER_FONT
        ws.cell(row=r, column=2).fill = HEADER_FILL
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=3, value="Description").font = HEADER_FONT
        ws.cell(row=r, column=3).fill = HEADER_FILL
        ws.cell(row=r, column=3).border = BORDER
        r += 1
        for level, desc in RATING_LEVELS:
            lc = ws.cell(row=r, column=2, value=level)
            lc.font = Font(name="Arial", size=10, bold=True, color="1F2E3D")
            lc.alignment = Alignment(horizontal="center")
            lc.border = BORDER
            dc = ws.cell(row=r, column=3, value=desc)
            dc.font = Font(name="Arial", size=10)
            dc.border = BORDER
            r += 1
        continue
    cell = ws.cell(row=r, column=2, value=text)
    cell.font = Font(name="Arial", size=11, bold=True, color="1F2E3D") if is_heading else Font(name="Arial", size=10.5)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if text and not is_heading:
        ws.row_dimensions[r].height = 26 if text.strip().startswith("•") else 30
    r += 1

# ============================================================= PROFILE
ws = wb.create_sheet("Profile")
headers = ["Name", "Position", "Section", "DateOfJoining", "OrgExperience (Years)", "OverallExperience (Years)"]
widths = [22, 28, 30, 16, 20, 22]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(headers, start=1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(headers))
ws.freeze_panes = "A2"

today = datetime.date.today()
SECTION_LABELS = [s["label"] for s in SECTIONS]

dv_section = DataValidation(type="list", formula1='"' + ",".join(SECTION_LABELS) + '"',
                             showErrorMessage=True, errorTitle="Invalid section",
                             error="Choose a level from the dropdown list")
ws.add_data_validation(dv_section)

dv_overall = DataValidation(type="decimal", operator="between", formula1=0, formula2=60,
                             showErrorMessage=True, errorTitle="Invalid entry",
                             error="Enter a number of years, e.g. 9.8")
ws.add_data_validation(dv_overall)

def write_profile_row(row, name, position, section_label, doj, overall, editable=True):
    ws.cell(row=row, column=1, value=name).font = BODY_FONT
    ws.cell(row=row, column=2, value=position).font = BODY_FONT
    sec_cell = ws.cell(row=row, column=3, value=section_label)
    sec_cell.font = BODY_FONT
    sec_cell.alignment = Alignment(horizontal="center")
    doj_cell = ws.cell(row=row, column=4, value=doj)
    doj_cell.font = BODY_FONT
    doj_cell.number_format = "DD-MMM-YYYY"
    org_cell = ws.cell(row=row, column=5, value=f"=ROUND((TODAY()-D{row})/365.25,1)")
    org_cell.font = BODY_FONT
    org_cell.fill = FORMULA_FILL
    org_cell.alignment = Alignment(horizontal="center")
    overall_cell = ws.cell(row=row, column=6, value=overall)
    overall_cell.font = BODY_FONT
    overall_cell.alignment = Alignment(horizontal="center")
    for c in range(1, 7):
        ws.cell(row=row, column=c).border = BORDER
    if editable:
        for c in (1, 2, 3, 4, 6):
            ws.cell(row=row, column=c).fill = INPUT_FILL
    dv_section.add(f"C{row}")
    dv_overall.add(f"F{row}")

row = 2
for s in SAMPLES:
    write_profile_row(row, s["name"], s["position"], SECTIONS[s["section"]]["label"], s["doj"], s["overall"])
    row += 1

# extra ready-to-fill blank rows with formula + section dropdown pre-built
for _ in range(15):
    write_profile_row(row, "", "", "", today, "")
    row += 1

# ============================================================= 5 LEVEL-SPECIFIC SKILLS SHEETS
dv_rating_registry = []
for s in SECTIONS:
    ws = wb.create_sheet(s["sheet"])
    skills = s["skills"]
    ws.column_dimensions["A"].width = 24
    for i in range(2, len(skills) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 15

    ws.cell(row=1, column=1, value="Name")
    for i, skill in enumerate(skills, start=2):
        ws.cell(row=1, column=i, value=skill)
    style_header(ws, 1, len(skills) + 1)
    ws.row_dimensions[1].height = 85
    ws.freeze_panes = "B2"

    dv_rating = DataValidation(type="whole", operator="between", formula1=0, formula2=5,
                                showErrorMessage=True, errorTitle="Invalid rating",
                                error="Rating must be a whole number from 0 to 5")
    ws.add_data_validation(dv_rating)

    def write_row(ws, row, name, ratings, dv, editable_name=True):
        name_cell = ws.cell(row=row, column=1, value=name)
        name_cell.font = BODY_FONT
        name_cell.border = BORDER
        if editable_name:
            name_cell.fill = INPUT_FILL
        for i, rating in enumerate(ratings, start=2):
            rc = ws.cell(row=row, column=i, value=rating)
            rc.font = BODY_FONT
            rc.alignment = Alignment(horizontal="center")
            rc.border = BORDER
            rc.fill = INPUT_FILL
            dv.add(rc.coordinate)

    r2 = 2
    sample = next((x for x in SAMPLES if x["section"] == SECTIONS.index(s)), None)
    if sample:
        write_row(ws, r2, sample["name"], sample["skills"], dv_rating)
        r2 += 1
    for _ in range(8):
        write_row(ws, r2, "", [0] * len(skills), dv_rating)
        r2 += 1

# ============================================================= SOFTWARE (common, wide format)
ws = wb.create_sheet("Software")
ws.column_dimensions["A"].width = 24
for i in range(2, len(SOFTWARE) + 2):
    ws.column_dimensions[get_column_letter(i)].width = 18
ws.cell(row=1, column=1, value="Name")
for i, sw in enumerate(SOFTWARE, start=2):
    ws.cell(row=1, column=i, value=sw)
style_header(ws, 1, len(SOFTWARE) + 1)
ws.row_dimensions[1].height = 45
ws.freeze_panes = "B2"

dv_rating_sw = DataValidation(type="whole", operator="between", formula1=0, formula2=5,
                               showErrorMessage=True, errorTitle="Invalid rating",
                               error="Rating must be a whole number from 0 to 5")
ws.add_data_validation(dv_rating_sw)

def write_sw_row(row, name, ratings, editable_name=True):
    name_cell = ws.cell(row=row, column=1, value=name)
    name_cell.font = BODY_FONT
    name_cell.border = BORDER
    if editable_name:
        name_cell.fill = INPUT_FILL
    for i, rating in enumerate(ratings, start=2):
        rc = ws.cell(row=row, column=i, value=rating)
        rc.font = BODY_FONT
        rc.alignment = Alignment(horizontal="center")
        rc.border = BORDER
        rc.fill = INPUT_FILL
        dv_rating_sw.add(rc.coordinate)

row = 2
for s in SAMPLES:
    write_sw_row(row, s["name"], s["sw"])
    row += 1
for _ in range(15):
    write_sw_row(row, "", [0] * len(SOFTWARE))
    row += 1

# reorder: Instructions, Profile, 5 skills sheets, Software
wb.move_sheet("Software", offset=-(len(wb.sheetnames) - wb.sheetnames.index("Software")) + 6)

wb.save("/home/claude/personnel_skills.xlsx")
print("saved. sheets:", wb.sheetnames)
